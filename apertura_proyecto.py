# -*- coding: utf-8 -*-
import io
import datetime
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import email.message
import email.utils
from pathlib import Path
import pandas as pd
import db_manager

def find_original_order_files(po, id_interno):
    """
    Localiza los archivos originales de la orden (correo .msg y PDF oficial)
    tanto en la base de datos po_archivos_adjuntos como en la carpeta data/correos/.
    """
    correos_dir = Path("data/correos")
    po_str = str(po).strip()
    po_nodash = po_str.replace('-', '')
    id_clean = str(id_interno).replace('INT-', '').strip()
    
    msg_bytes, msg_name = None, None
    pdf_bytes, pdf_name = None, None
    
    # 1. Cabecera DB
    try:
        df_cab, _ = db_manager.get_po_by_folio(po_str)
        if df_cab.empty:
            df_cab, _ = db_manager.get_po_by_folio(po_nodash)
        if not df_cab.empty:
            r = df_cab.iloc[0]
            if r.get('archivo_correo'):
                p = correos_dir / str(r['archivo_correo'])
                if p.exists():
                    msg_name = p.name
                    with open(p, 'rb') as f:
                        msg_bytes = f.read()
            if r.get('archivo_pdf'):
                p = correos_dir / str(r['archivo_pdf'])
                if p.exists():
                    pdf_name = p.name
                    with open(p, 'rb') as f:
                        pdf_bytes = f.read()
    except Exception as e:
        print(f"[WARN] find_original_order_files cabecera: {e}")

    # 2. po_archivos_adjuntos
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre_archivo, tipo FROM po_archivos_adjuntos
            WHERE po IN (?, ?) OR id_interno = ?
        """, (po_str, po_nodash, str(id_interno)))
        for f_id, fn, ft in cur.fetchall():
            if str(ft).lower() == 'msg' and not msg_bytes:
                _, _, b = db_manager.get_contenido_archivo_por_id(f_id)
                if b:
                    msg_name, msg_bytes = fn, b
            elif str(ft).lower() == 'pdf' and not pdf_bytes:
                _, _, b = db_manager.get_contenido_archivo_por_id(f_id)
                if b:
                    pdf_name, pdf_bytes = fn, b
        conn.close()
    except Exception as e:
        print(f"[WARN] find_original_order_files po_archivos_adjuntos: {e}")

    # 3. Disco data/correos/
    if correos_dir.exists():
        for f in sorted(correos_dir.iterdir()):
            fn = f.name
            low = fn.lower()
            m_po = (po_str in fn) or (po_nodash in fn)
            m_id = (f"INT {id_clean}" in fn) or (f"INT_{id_clean}" in fn) or (f"INT{id_clean}" in fn) or (f"INT 0{id_clean}" in fn)
            
            if low.endswith('.msg') and not msg_bytes and (m_po or m_id):
                msg_name = fn
                try:
                    with open(f, 'rb') as fp:
                        msg_bytes = fp.read()
                except Exception:
                    pass
            if low.endswith('.pdf') and not pdf_bytes and (m_po or m_id):
                pdf_name = fn
                try:
                    with open(f, 'rb') as fp:
                        pdf_bytes = fp.read()
                except Exception:
                    pass

    return msg_bytes, msg_name, pdf_bytes, pdf_name

def _extract_val(row, keys, default=""):
    for k in keys:
        if k in row:
            val = row.get(k)
            if val is not None and not pd.isna(val):
                s_val = str(val).strip()
                if s_val and s_val.lower() not in ('nan', 'none', 'null'):
                    return s_val
    return default

def _ensure_partidas_skus(po, df_partidas):
    """
    Garantiza que el DataFrame de partidas contenga siempre los campos sku_cliente
    y clave_sku completos, consultando directamente la base de datos po_partidas de SQLite
    en caso de que vengan vacíos o nulos.
    """
    po_str = str(po).strip()
    po_nodash = po_str.replace('-', '')
    po_dash = f"{po_nodash[:4]}-{po_nodash[4:]}" if len(po_nodash) == 8 else po_str
    
    conn = db_manager.get_connection()
    df_db_parts = pd.read_sql_query(
        """SELECT item_no, sku_cliente, clave_sku, descripcion_producto, 
                  cantidad_requerida, unidad, precio_unitario, precio_total, 
                  fecha_entrega, parcialidad, observaciones_partida 
           FROM po_partidas 
           WHERE po = ? OR po = ? OR po = ?
           ORDER BY item_no ASC""",
        conn, params=[po_str, po_nodash, po_dash]
    )
    conn.close()

    if df_partidas is None or df_partidas.empty:
        return df_db_parts.copy()

    df_out = df_partidas.copy()
    if not df_db_parts.empty:
        # Mapas directos por item_no
        map_cli = {}
        map_planta = {}
        for _, r_db in df_db_parts.iterrows():
            it_key = str(r_db['item_no']).strip()
            map_cli[it_key] = str(r_db['sku_cliente'] or '').strip()
            map_planta[it_key] = str(r_db['clave_sku'] or '').strip()

        # Enriquecer o rellenar sku_cliente
        if 'sku_cliente' not in df_out.columns:
            df_out['sku_cliente'] = df_out['item_no'].astype(str).str.strip().map(map_cli).fillna('')
        else:
            def _fill_c(row):
                v = str(row.get('sku_cliente', '') or '').strip()
                if not v or v.lower() in ('nan', 'none', 'null'):
                    it = str(row.get('item_no', '')).strip()
                    return map_cli.get(it, '')
                return v
            df_out['sku_cliente'] = df_out.apply(_fill_c, axis=1)

        # Enriquecer o rellenar clave_sku
        if 'clave_sku' not in df_out.columns:
            df_out['clave_sku'] = df_out['item_no'].astype(str).str.strip().map(map_planta).fillna('')
        else:
            def _fill_p(row):
                v = str(row.get('clave_sku', '') or '').strip()
                if not v or v.lower() in ('nan', 'none', 'null'):
                    it = str(row.get('item_no', '')).strip()
                    return map_planta.get(it, '')
                return v
            df_out['clave_sku'] = df_out.apply(_fill_p, axis=1)

    return df_out

def generate_apertura_piezas_excel(po, id_interno, cab_info, df_partidas):
    """
    Genera el archivo oficial de Apertura de Proyecto / Lista de Piezas en formato .xlsx
    con diseño corporativo formal SIGRAMA, datos de cabecera y desglose completo de despiece.
    """
    df_partidas = _ensure_partidas_skus(po, df_partidas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista_de_Piezas"
    ws.views.sheetView[0].showGridLines = True

    C_SLATE_DARK = "0F172A"
    C_SLATE_MID  = "1E293B"
    C_RED_SIG    = "EC2024"
    C_BORDER_THIN = Side(border_style="thin", color="CBD5E1")
    C_BORDER_MED  = Side(border_style="medium", color="0F172A")
    C_BORDER_DBL  = Side(border_style="double", color="0F172A")

    thin_border = Border(top=C_BORDER_THIN, bottom=C_BORDER_THIN, left=C_BORDER_THIN, right=C_BORDER_THIN)
    total_border = Border(top=C_BORDER_THIN, bottom=C_BORDER_DBL, left=C_BORDER_THIN, right=C_BORDER_THIN)

    # 1. Header Banner
    ws.merge_cells("A1:K1")
    c1 = ws["A1"]
    c1.value = "INDUSTRIA SIGRAMA S.A. DE C.V.  —  LISTA DE PIEZAS / APERTURA DE PROYECTO"
    c1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c1.fill = PatternFill(start_color=C_SLATE_DARK, end_color=C_SLATE_DARK, fill_type="solid")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells("A2:K2")
    c2 = ws["A2"]
    c2.value = f"Documento Oficial de Despiece y Programación Operativa | Emisión: {now_str} | Control de Planta y Almacén"
    c2.font = Font(name="Calibri", size=9.5, italic=True, color="94A3B8")
    c2.fill = PatternFill(start_color=C_SLATE_MID, end_color=C_SLATE_MID, fill_type="solid")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6

    # 2. Tarjetas de Metadatos del Proyecto
    def set_meta_cell(cell_ref, label, val):
        ws[cell_ref].value = f"{label}: {val}"
        ws[cell_ref].font = Font(name="Calibri", size=9.5, bold=True, color=C_SLATE_DARK)
        ws[cell_ref].fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:C4")
    set_meta_cell("A4", "No. Proyecto Interno", str(id_interno if id_interno else "S/N"))
    ws.merge_cells("D4:F4")
    set_meta_cell("D4", "Orden de Compra (PO)", str(po))
    ws.merge_cells("G4:I4")
    set_meta_cell("G4", "Proyecto", str(cab_info.get('proyecto', 'N/A')))
    ws.merge_cells("J4:K4")
    set_meta_cell("J4", "Estatus", str(cab_info.get('estatus_general', 'Registrada')))

    ws.merge_cells("A5:C5")
    set_meta_cell("A5", "Comprador", str(cab_info.get('comprador', 'N/A')))
    ws.merge_cells("D5:F5")
    set_meta_cell("D5", "Solicitante", str(cab_info.get('solicitante', 'N/A')))
    ws.merge_cells("G5:I5")
    set_meta_cell("G5", "Fecha Llegada PO", str(cab_info.get('fecha_llegada', 'N/A')))
    ws.merge_cells("J5:K5")
    set_meta_cell("J5", "Fecha Entrega Req.", str(cab_info.get('fecha_solicitada', 'N/A')))

    for r_idx in range(4, 6):
        ws.row_dimensions[r_idx].height = 22
        for col_idx in range(1, 12):
            ws.cell(row=r_idx, column=col_idx).border = thin_border

    ws.row_dimensions[6].height = 8

    # 3. Encabezados de Columnas
    headers = [
        ("#", 6, "center"),
        ("SKU Cliente", 18, "center"),
        ("SKU Planta (Clave)", 20, "center"),
        ("Descripción del Producto", 38, "left"),
        ("Cant. Requerida", 16, "right"),
        ("Unidad", 10, "center"),
        ("P. Unitario ($)", 15, "right"),
        ("Importe Total ($)", 18, "right"),
        ("Fecha Entrega", 15, "center"),
        ("Parcialidad", 13, "center"),
        ("Observaciones / Notas", 25, "left")
    ]

    header_row = 7
    ws.row_dimensions[header_row].height = 24
    for idx, (h_title, w, h_align) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx)
        cell.value = h_title
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=C_SLATE_DARK, end_color=C_SLATE_DARK, fill_type="solid")
        cell.alignment = Alignment(horizontal=h_align, vertical="center")
        cell.border = Border(top=C_BORDER_MED, bottom=Side(border_style="medium", color=C_RED_SIG), left=C_BORDER_THIN, right=C_BORDER_THIN)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = w

    # 4. Filas de Partidas
    curr_row = 8
    df_sorted = df_partidas.copy()
    if 'item_no' in df_sorted.columns:
        df_sorted['item_no_num'] = pd.to_numeric(df_sorted['item_no'], errors='coerce').fillna(9999)
        df_sorted = df_sorted.sort_values('item_no_num')

    for _, row in df_sorted.iterrows():
        ws.row_dimensions[curr_row].height = 20
        bg_color = "FFFFFF" if curr_row % 2 == 0 else "F8FAFC"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        i_no = _extract_val(row, ['item_no'], str(curr_row - 7))
        sk_c = _extract_val(row, ['sku_cliente', 'SKU Cliente', 'sku_cli', 'SKU_Cliente', 'SKU'])
        sk_p = _extract_val(row, ['clave_sku', 'SKU Planta', 'clave', 'SKU_Planta', 'Clave SKU'])
        desc = _extract_val(row, ['descripcion_producto', 'Descripcion', 'descripcion', 'Descripción'])
        cant = float(_extract_val(row, ['cantidad_requerida', 'Cantidad', 'cant'], 0) or 0)
        unid = _extract_val(row, ['unidad', 'Unidad'], 'PZA').upper()
        pu   = float(_extract_val(row, ['precio_unitario', 'P. Unitario'], 0) or 0)
        pt   = float(_extract_val(row, ['precio_total', 'Importe Total'], cant * pu) or (cant * pu))
        fe   = _extract_val(row, ['fecha_entrega', 'Fecha_Entrega', 'Fecha Entrega'])
        parc = _extract_val(row, ['parcialidad', 'Parcialidad'], 'P1')
        obs  = _extract_val(row, ['observaciones_partida', 'estatus_partida_360', 'Observaciones'])

        values = [
            (i_no, "center", "@"),
            (sk_c, "center", "@"),
            (sk_p, "center", "@"),
            (desc, "left", "@"),
            (cant, "right", '#,##0 "pzas"'),
            (unid, "center", "@"),
            (pu, "right", '"$"#,##0.00'),
            (pt, "right", '"$"#,##0.00'),
            (fe, "center", "@"),
            (parc, "center", "@"),
            (obs, "left", "@")
        ]

        for col_idx, (val, align, num_fmt) in enumerate(values, start=1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.value = val
            cell.font = Font(name="Calibri", size=9.5)
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.fill = row_fill
            cell.border = thin_border
            if num_fmt != "@":
                cell.number_format = num_fmt

        curr_row += 1

    # 5. Fila de Totales Generales
    ws.row_dimensions[curr_row].height = 24
    ws.merge_cells(f"A{curr_row}:D{curr_row}")
    lbl_tot = ws[f"A{curr_row}"]
    lbl_tot.value = "TOTAL GENERAL DE PIEZAS E IMPORTE"
    lbl_tot.font = Font(name="Calibri", size=10, bold=True, color=C_SLATE_DARK)
    lbl_tot.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 5):
        ws.cell(row=curr_row, column=col_idx).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws.cell(row=curr_row, column=col_idx).border = total_border

    c_tot_cant = ws.cell(row=curr_row, column=5)
    c_tot_cant.value = f"=SUM(E8:E{curr_row-1})"
    c_tot_cant.number_format = '#,##0 "pzas"'
    c_tot_cant.font = Font(name="Calibri", size=10, bold=True, color=C_SLATE_DARK)
    c_tot_cant.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    c_tot_cant.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_cant.border = total_border

    for col_idx in range(6, 8):
        c_b = ws.cell(row=curr_row, column=col_idx)
        c_b.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_b.border = total_border

    c_tot_imp = ws.cell(row=curr_row, column=8)
    c_tot_imp.value = f"=SUM(H8:H{curr_row-1})"
    c_tot_imp.number_format = '"$"#,##0.00'
    c_tot_imp.font = Font(name="Calibri", size=10, bold=True, color=C_SLATE_DARK)
    c_tot_imp.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    c_tot_imp.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_imp.border = total_border

    for col_idx in range(9, 12):
        c_b = ws.cell(row=curr_row, column=col_idx)
        c_b.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_b.border = total_border

    ws.auto_filter.ref = f"A7:K{curr_row-1}"
    ws.freeze_panes = "A8"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_apertura_eml(po, id_interno, cab_info, df_partidas, msg_bytes=None, msg_name=None, pdf_bytes=None, pdf_name=None, excel_bytes=None):
    """
    Genera el correo formal de Apertura de Proyecto Interno en formato RFC 822 (.eml)
    100% COMPATIBLE CON MICROSOFT OUTLOOK (Word Engine):
    - Usa tablas HTML puras con bgcolor="#18181B" para que Outlook jamás deje el fondo en blanco.
    - Encabezado con banner idéntico a la Ficha de Trazabilidad 360° (segunda imagen).
    - Codificación Content-Transfer-Encoding en base64 para evitar saltos suaves de línea con '='.
    - Adjuntos embebidos: Lista de Piezas (.xlsx), Correo Original (.msg) y PDF de la PO.
    """
    df_partidas = _ensure_partidas_skus(po, df_partidas)

    msg = email.message.EmailMessage()

    def _clean(v, d=""):
        if v is None or pd.isna(v):
            return d
        s = str(v).replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s if s else d

    po_clean = _clean(po, "S/N")
    id_clean = _clean(id_interno, "INT-S/N")
    prj_clean = _clean(cab_info.get('proyecto', 'PROYECTO SIGRAMA'), 'PROYECTO SIGRAMA')
    cmp_clean = _clean(cab_info.get('comprador', 'Compras'), 'Compras')
    sol_clean = _clean(cab_info.get('solicitante', 'Solicitante'), 'Solicitante')
    f_lleg = _clean(cab_info.get('fecha_llegada', 'N/A'), 'N/A')
    f_sol = _clean(cab_info.get('fecha_solicitada', 'N/A'), 'N/A')
    est_gen = _clean(cab_info.get('estatus_general', 'Registrada (En Espera)'), 'Registrada (En Espera)')

    # Determinar color de estatus
    est_low = est_gen.lower()
    if 'cancel' in est_low:
        est_bg = "#EF4444"
        est_badge = "🚫 CANCELADA"
    elif 'total' in est_low or '100' in est_low:
        est_bg = "#10B981"
        est_badge = "● Remisionada Total"
    elif 'parcial' in est_low:
        est_bg = "#3B82F6"
        est_badge = "● Remisión Parcial"
    elif 'fab' in est_low or 'proceso' in est_low:
        est_bg = "#F59E0B"
        est_badge = "● En Fabricación"
    else:
        est_bg = "#475569"
        est_badge = "● Registrada (En Espera)"

    tot_pzas = float(df_partidas['cantidad_requerida'].sum()) if 'cantidad_requerida' in df_partidas.columns else 0.0
    tot_imp = float(cab_info.get('total', 0) or 0)
    if tot_imp == 0.0 and 'precio_total' in df_partidas.columns:
        tot_imp = float(df_partidas['precio_total'].sum())

    msg['Subject'] = f"[APERTURA DE PROYECTO INTERNO] {id_clean} - OC {po_clean} | PROYECTO: {prj_clean}"
    msg['From'] = 'operaciones@sigrama.com.mx'
    msg['To'] = f"{cmp_clean.lower().replace(' ', '.')}@sigrama.com.mx"
    msg['Cc'] = 'operaciones@sigrama.com.mx, produccion@sigrama.com.mx, calidad@sigrama.com.mx, almacen@sigrama.com.mx'
    msg['Date'] = email.utils.formatdate(localtime=True)

    # Filas de la tabla de partidas
    filas_html = ""
    for idx, (_, r) in enumerate(df_partidas.head(150).iterrows(), start=1):
        bg_col = "#FFFFFF" if idx % 2 != 0 else "#F8FAFC"
        i_no = _extract_val(r, ['item_no'], str(idx))
        sk_c = _extract_val(r, ['sku_cliente', 'SKU Cliente', 'sku_cli', 'SKU_Cliente', 'SKU'])
        sk_p = _extract_val(r, ['clave_sku', 'SKU Planta', 'clave', 'SKU_Planta', 'Clave SKU'])
        desc = _extract_val(r, ['descripcion_producto', 'Descripcion', 'descripcion', 'Descripción'])
        cant = float(_extract_val(r, ['cantidad_requerida', 'Cantidad', 'cant'], 0) or 0)
        unid = _extract_val(r, ['unidad', 'Unidad'], 'PZA').upper()
        fe   = _extract_val(r, ['fecha_entrega', 'Fecha_Entrega', 'Fecha Entrega'])

        filas_html += f"""
        <tr bgcolor="{bg_col}" style="background-color: {bg_col}; border-bottom: 1px solid #E2E8F0;">
            <td align="center" style="padding: 8px 6px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #475569; border-bottom: 1px solid #E2E8F0;">{i_no}</td>
            <td align="left" style="padding: 8px 8px; font-family: Arial, sans-serif; font-size: 12px; font-weight: bold; color: #0F172A; white-space: nowrap; border-bottom: 1px solid #E2E8F0;">{sk_c}</td>
            <td align="left" style="padding: 8px 8px; font-family: Arial, sans-serif; font-size: 12px; font-weight: bold; color: #2563EB; white-space: nowrap; border-bottom: 1px solid #E2E8F0;">{sk_p}</td>
            <td align="left" style="padding: 8px 8px; font-family: Arial, sans-serif; font-size: 11.5px; color: #334155; border-bottom: 1px solid #E2E8F0;">{desc}</td>
            <td align="right" style="padding: 8px 8px; font-family: Arial, sans-serif; font-size: 12px; font-weight: bold; color: #0F172A; border-bottom: 1px solid #E2E8F0;">{cant:,.0f}</td>
            <td align="center" style="padding: 8px 6px; font-family: Arial, sans-serif; font-size: 11.5px; color: #64748B; border-bottom: 1px solid #E2E8F0;">{unid}</td>
            <td align="center" style="padding: 8px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #059669; border-bottom: 1px solid #E2E8F0;">{fe}</td>
        </tr>
        """

    mas_filas_nota = ""
    if len(df_partidas) > 150:
        mas_filas_nota = f"""
        <tr>
            <td colspan="7" bgcolor="#FEF3C7" style="padding: 12px; text-align: center; background-color: #FEF3C7; color: #B45309; font-weight: bold; font-size: 11.5px; font-family: Arial, sans-serif;">
                ⚠️ Mostrando 150 de {len(df_partidas)} partidas. Consulte la lista completa en el archivo Excel adjunto: Lista_Piezas_Despiece_{id_clean}_{po_clean}.xlsx
            </td>
        </tr>
        """

    adjuntos_html = []
    adjuntos_html.append(f"<li>📊 <b>Lista de Piezas Oficial:</b> <code>Lista_Piezas_Despiece_{id_clean}_{po_clean}.xlsx</code></li>")
    if msg_name:
        adjuntos_html.append(f"<li>📧 <b>Correo Original Embebido:</b> <code>{msg_name}</code></li>")
    if pdf_name:
        adjuntos_html.append(f"<li>📄 <b>Orden de Compra Oficial (PDF):</b> <code>{pdf_name}</code></li>")
    adjuntos_str = "".join(adjuntos_html)

    # HTML 100% COMPATIBLE CON OUTLOOK (Tablas anidadas con bgcolor en TD)
    html_content = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Apertura Oficial de Proyecto Interno</title>
</head>
<body bgcolor="#F1F5F9" style="background-color: #F1F5F9; margin: 0; padding: 0; font-family: Arial, Helvetica, sans-serif;">

<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#F1F5F9" style="background-color: #F1F5F9; padding: 25px 0;">
  <tr>
    <td align="center" style="padding: 0 10px;">
      
      <!-- CONTENEDOR PRINCIPAL (Ancho fijo compatible con Outlook) -->
      <table width="820" cellpadding="0" cellspacing="0" border="0" bgcolor="#FFFFFF" style="width: 820px; max-width: 820px; background-color: #FFFFFF; border: 1px solid #CBD5E1; border-radius: 10px; overflow: hidden;">
        
        <!-- ── BANNER OSCURO SUPERIOR IDÉNTICO A LA FICHA 360° (IMAGEN 2) ────────────────── -->
        <tr>
          <td bgcolor="#18181B" style="background-color: #18181B; border-left: 8px solid #EC2024; padding: 22px 28px;">
            
            <!-- Fila superior de membrete -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom: 8px;">
              <tr>
                <td align="left">
                  <span style="color: #EC2024; font-size: 11.5px; font-weight: bold; letter-spacing: 1.5px; text-transform: uppercase; font-family: Arial, sans-serif;">
                    INDUSTRIA SIGRAMA S.A. DE C.V. &nbsp;—&nbsp; APERTURA OFICIAL DE PROYECTO INTERNO
                  </span>
                </td>
              </tr>
            </table>

            <!-- Fila central: Badge Rojo + ORDEN DE COMPRA + Estatus -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom: 14px;">
              <tr>
                <td align="left" style="vertical-align: middle;">
                  <table cellpadding="0" cellspacing="0" border="0">
                    <tr>
                      <td bgcolor="#EC2024" style="background-color: #EC2024; border-radius: 6px; padding: 6px 16px; text-align: center;">
                        <span style="color: #FFFFFF; font-size: 24px; font-weight: 900; letter-spacing: 1px; font-family: Arial, sans-serif; display: inline-block;">
                          {id_clean}
                        </span>
                      </td>
                      <td style="padding-left: 14px; vertical-align: middle;">
                        <span style="color: #FFFFFF; font-size: 26px; font-weight: 900; letter-spacing: -0.5px; font-family: Arial, sans-serif;">
                          ORDEN DE COMPRA: <span style="color: #EC2024;">{po_clean}</span>
                        </span>
                      </td>
                    </tr>
                  </table>
                </td>
                <td align="right" style="vertical-align: middle;">
                  <table cellpadding="0" cellspacing="0" border="0">
                    <tr>
                      <td bgcolor="{est_bg}" style="background-color: {est_bg}; border-radius: 20px; padding: 7px 18px; text-align: center;">
                        <span style="color: #FFFFFF; font-size: 13px; font-weight: bold; font-family: Arial, sans-serif; white-space: nowrap;">
                          {est_badge}
                        </span>
                      </td>
                    </tr>
                  </table>
                </td>
              </tr>
            </table>

            <!-- Fila inferior: Metadatos con iconos idéntico a la segunda imagen -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-top: 1px solid #334155; padding-top: 10px;">
              <tr>
                <td style="padding-top: 10px; font-family: Arial, sans-serif; font-size: 13.5px; color: #E5E7EB; line-height: 1.5;">
                  <span style="margin-right: 12px;">🏗️ Proyecto: <b style="color: #FFFFFF; font-size: 14.5px;">{prj_clean}</b></span>
                  <span style="color: #64748B; margin-right: 12px;">|</span>
                  <span style="margin-right: 12px;">👤 Solicitante: <b style="color: #FFFFFF;">{sol_clean}</b></span>
                  <span style="color: #64748B; margin-right: 12px;">|</span>
                  <span>💼 Comprador: <b style="color: #FFFFFF;">{cmp_clean}</b></span>
                </td>
              </tr>
              <tr>
                <td style="padding-top: 6px; font-family: Arial, sans-serif; font-size: 12px; color: #94A3B8;">
                  <span style="margin-right: 12px;">📅 Llegada PO: <b style="color: #E2E8F0;">{f_lleg}</b></span>
                  <span style="color: #64748B; margin-right: 12px;">|</span>
                  <span style="margin-right: 12px;">🎯 Entrega Solicitada: <b style="color: #F87171;">{f_sol}</b></span>
                  <span style="color: #64748B; margin-right: 12px;">|</span>
                  <span style="margin-right: 12px;">📦 Total Piezas: <b style="color: #FFFFFF;">{tot_pzas:,.0f} pzas</b></span>
                  <span style="color: #64748B; margin-right: 12px;">|</span>
                  <span>💰 Importe Total: <b style="color: #34D399;">${tot_imp:,.2f} MXN</b></span>
                </td>
              </tr>
            </table>

          </td>
        </tr>

        <!-- ── CUERPO DEL CORREO ──────────────────────────────────────────────────────────── -->
        <tr>
          <td style="padding: 24px 28px;">
            <p style="font-family: Arial, sans-serif; font-size: 13.5px; color: #334155; line-height: 1.6; margin-top: 0;">
              Estimado equipo operativo, compras y jefatura de planta:<br>
              Se ha formalizado la <b>Apertura de Proyecto Interno</b> para la Orden de Compra <b>{po_clean}</b>. A continuación se detallan las especificaciones, fechas solicitadas y despiece de piezas para el arranque inmediato en talleres y almacén.
            </p>

            <!-- TITULO TABLA PARTIDAS -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin: 20px 0 10px 0;">
              <tr>
                <td>
                  <span style="color: #0F172A; font-family: Arial, sans-serif; font-size: 13.5px; font-weight: bold; text-transform: uppercase; letter-spacing: 0.5px;">
                    📑 RESUMEN DE PIEZAS / PARTIDAS REQUERIDAS ({len(df_partidas)} PARTIDAS)
                  </span>
                </td>
              </tr>
            </table>

            <!-- TABLA DE PARTIDAS -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse: collapse; border: 1px solid #CBD5E1; margin-bottom: 22px;">
              <thead>
                <tr bgcolor="#0F172A" style="background-color: #0F172A; color: #FFFFFF;">
                  <th bgcolor="#0F172A" width="40" align="center" style="padding: 9px 6px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">#</th>
                  <th bgcolor="#0F172A" width="115" align="left" style="padding: 9px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">SKU Cliente</th>
                  <th bgcolor="#0F172A" width="125" align="left" style="padding: 9px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">SKU Planta</th>
                  <th bgcolor="#0F172A" align="left" style="padding: 9px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">Descripción</th>
                  <th bgcolor="#0F172A" width="85" align="right" style="padding: 9px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">Cant. Req.</th>
                  <th bgcolor="#0F172A" width="55" align="center" style="padding: 9px 6px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">Unidad</th>
                  <th bgcolor="#0F172A" width="85" align="center" style="padding: 9px 8px; font-family: Arial, sans-serif; font-size: 11.5px; font-weight: bold; color: #FFFFFF; border-bottom: 2px solid #EC2024;">F. Entrega</th>
                </tr>
              </thead>
              <tbody>
                {filas_html}
                {mas_filas_nota}
              </tbody>
            </table>

            <!-- CAJA DE PLAN DE ACCION OPERATIVO -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#EFF6FF" style="background-color: #EFF6FF; border-left: 4px solid #3B82F6; border-radius: 6px; margin-bottom: 20px;">
              <tr>
                <td style="padding: 14px 18px; font-family: Arial, sans-serif;">
                  <b style="color: #1E40AF; font-size: 12.5px; display: block; margin-bottom: 6px;">⚙️ Plan de Acción Operativo Inmediato:</b>
                  <ul style="margin: 0; padding-left: 20px; font-size: 12px; color: #1E3A8A; line-height: 1.6;">
                    <li><b>Corte y Doblez (Nesting):</b> Generar nidos Pronest y programar Órdenes de Fabricación (OFs) conforme al despiece adjunto.</li>
                    <li><b>Ensamble & Soldadura:</b> Coordinar ensambles y soldadura cumpliendo con fechas prometidas por parcialidad.</li>
                    <li><b>Almacén & Embarques:</b> Identificación clara por tarima y generación oportuna de Remisiones.</li>
                  </ul>
                </td>
              </tr>
            </table>

            <!-- CAJA DE ADJUNTOS EMBEBIDOS -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#F8FAFC" style="background-color: #F8FAFC; border: 1px solid #CBD5E1; border-radius: 6px;">
              <tr>
                <td style="padding: 14px 18px; font-family: Arial, sans-serif;">
                  <b style="color: #0F172A; font-size: 12.5px; display: block; margin-bottom: 6px;">📎 Documentos Oficiales Embebidos en este Correo:</b>
                  <ul style="margin: 0; padding-left: 20px; font-size: 12px; color: #334155; line-height: 1.6;">
                    {adjuntos_str}
                  </ul>
                </td>
              </tr>
            </table>

          </td>
        </tr>

        <!-- ── PIE DE PÁGINA ────────────────────────────────────────────────────────────── -->
        <tr>
          <td bgcolor="#F8FAFC" style="background-color: #F8FAFC; border-top: 1px solid #E2E8F0; padding: 14px 28px; text-align: center; font-family: Arial, sans-serif; font-size: 11px; color: #94A3B8;">
            Industria Sigrama S.A. de C.V. | Sistema Integral de Seguimiento y Trazabilidad 360° | Torreón, Coahuila
          </td>
        </tr>

      </table>

    </td>
  </tr>
</table>

</body>
</html>
"""

    msg.set_content(f"Apertura de Proyecto Interno {id_clean} - OC {po_clean}. Consulte la versión HTML y los archivos adjuntos.")
    # Usar CTE base64 para evitar saltos de línea Quoted-Printable que corrompen palabras con '='
    msg.add_alternative(html_content, subtype='html', cte='base64')

    # Adjuntos
    if excel_bytes:
        msg.add_attachment(
            excel_bytes,
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"Lista_Piezas_Despiece_{id_clean}_{po_clean}.xlsx"
        )
    if msg_bytes and msg_name:
        msg.add_attachment(
            msg_bytes,
            maintype='application',
            subtype='vnd.ms-outlook',
            filename=msg_name
        )
    if pdf_bytes and pdf_name:
        msg.add_attachment(
            pdf_bytes,
            maintype='application',
            subtype='pdf',
            filename=pdf_name
        )

    return msg.as_bytes()
