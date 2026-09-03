import sys
sys.stdout.reconfigure(encoding='utf-8')
import io
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from db_manager import get_all_pos, get_all_partidas
from remisiones_sync import get_global_pos_tracking_summary

def build_executive_excel(df_data, df_partidas=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz_Ordenes_360"
    ws.views.sheetView[0].showGridLines = True
    
    # Colores corporativos SIGRAMA
    C_SLATE_DARK = "0F172A"
    C_SLATE_MID  = "1E293B"
    C_RED_SIG    = "EC2024"
    C_BLUE_FAB   = "1E3A8A"
    C_AMBER_ENT  = "78350F"
    C_GREEN_REM  = "064E3B"
    C_RED_PEN    = "7C2D12"
    
    # ── 1. Banner Principal (Fila 1 y 2) ──────────────────────────────────────
    ws.merge_cells("A1:O1")
    cell_t1 = ws["A1"]
    cell_t1.value = "INDUSTRIA SIGRAMA S.A. DE C.V.  —  MATRIZ DE CONTROL 360° DE ÓRDENES DE COMPRA"
    cell_t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    cell_t1.fill = PatternFill(start_color=C_SLATE_DARK, end_color=C_SLATE_DARK, fill_type="solid")
    cell_t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    
    # Fecha y metadata
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    tot_pos = len(df_data)
    tot_req = float(df_data['piezas_requeridas'].sum()) if 'piezas_requeridas' in df_data.columns else 0
    tot_fab = float(df_data['piezas_fabricadas'].sum()) if 'piezas_fabricadas' in df_data.columns else 0
    tot_ent = float(df_data['piezas_entarimadas'].sum()) if 'piezas_entarimadas' in df_data.columns else 0
    tot_rem = float(df_data['piezas_remisionadas'].sum()) if 'piezas_remisionadas' in df_data.columns else 0
    tot_pen = float(df_data['piezas_pendientes'].sum()) if 'piezas_pendientes' in df_data.columns else 0
    tot_imp = float(df_data['total'].sum()) if 'total' in df_data.columns else 0
    pct_glob = (tot_rem / tot_req * 100.0) if tot_req > 0 else 0.0
    
    ws.merge_cells("A2:O2")
    cell_t2 = ws["A2"]
    cell_t2.value = f"Reporte Oficial de Cadena de Suministro | Emisión: {now_str} | {tot_pos} Órdenes Activas | Cumplimiento Global: {pct_glob:.1f}% | Importe Total: ${tot_imp:,.2f} MXN"
    cell_t2.font = Font(name="Calibri", size=9.5, italic=True, color="94A3B8")
    cell_t2.fill = PatternFill(start_color=C_SLATE_MID, end_color=C_SLATE_MID, fill_type="solid")
    cell_t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # ── 2. Tarjetas de KPI Resumen (Fila 4) ───────────────────────────────────
    ws.row_dimensions[3].height = 6
    
    kpis = [
        ("A4:C4", f"1. REQUERIDAS: {tot_req:,.0f} pzas", "0F172A", "F8FAFC", "0F172A"),
        ("D4:F4", f"2. FABRICADAS: {tot_fab:,.0f} pzas", "1D4ED8", "EFF6FF", "3B82F6"),
        ("G4:I4", f"3. ENTARIMADAS: {tot_ent:,.0f} pzas", "B45309", "FEF3C7", "F59E0B"),
        ("J4:L4", f"4. REMISIONADAS: {tot_rem:,.0f} pzas", "15803D", "DCFCE7", "10B981"),
        ("M4:O4", f"5. PENDIENTES: {tot_pen:,.0f} pzas", "B91C1C", "FEE2E2", "EF4444"),
    ]
    for rng, text, fg, bg, border_c in kpis:
        ws.merge_cells(rng)
        first_c = ws[rng.split(":")[0]]
        first_c.value = text
        first_c.font = Font(name="Calibri", size=10, bold=True, color=fg)
        first_c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        first_c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Borde para la celda unificada
        thin_side = Side(border_style="medium", color=border_c)
        for row_c in ws[rng]:
            for cell in row_c:
                cell.border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 6
    
    # ── 3. Encabezados de Columnas (Fila 6) ───────────────────────────────────
    headers = [
        ("ID Interno", C_SLATE_DARK, "center", 12),
        ("PO / Folio", C_SLATE_DARK, "center", 14),
        ("Proyecto", C_SLATE_DARK, "left", 16),
        ("Fecha Llegada", C_SLATE_DARK, "center", 14),
        ("Part. #", C_SLATE_DARK, "center", 10),
        ("1. Req. (PO)", "1E293B", "right", 15),
        ("🔵 2. Fabricadas", C_BLUE_FAB, "right", 16),
        ("📦 3. Entarimadas", C_AMBER_ENT, "right", 16),
        ("🟢 4. Remisionadas", C_GREEN_REM, "right", 16),
        ("⏳ 5. Pendientes", C_RED_PEN, "right", 16),
        ("% Cumplimiento", C_SLATE_DARK, "right", 16),
        ("Estatus Entrega", C_SLATE_DARK, "center", 24),
        ("Importe Total ($)", C_SLATE_DARK, "right", 18),
        ("Comprador", C_SLATE_DARK, "left", 22),
        ("Solicitante", C_SLATE_DARK, "left", 22),
    ]
    
    hdr_row = 6
    ws.row_dimensions[hdr_row].height = 26
    border_red_bottom = Border(
        bottom=Side(border_style="medium", color=C_RED_SIG),
        left=Side(border_style="thin", color="334155"),
        right=Side(border_style="thin", color="334155")
    )
    
    for col_idx, (col_name, bg_c, align_h, col_w) in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.value = col_name
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=bg_c, end_color=bg_c, fill_type="solid")
        cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)
        cell.border = border_red_bottom
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_w
        
    # ── 4. Relleno de Datos (Filas 7 en adelante) ─────────────────────────────
    border_data = Border(
        top=Side(border_style="thin", color="E2E8F0"),
        bottom=Side(border_style="thin", color="E2E8F0"),
        left=Side(border_style="thin", color="E2E8F0"),
        right=Side(border_style="thin", color="E2E8F0")
    )
    
    start_row = 7
    for i, (_, r) in enumerate(df_data.iterrows()):
        curr_row = start_row + i
        ws.row_dimensions[curr_row].height = 20
        zebra_bg = "F8FAFC" if (i % 2 == 1) else "FFFFFF"
        fill_zebra = PatternFill(start_color=zebra_bg, end_color=zebra_bg, fill_type="solid")
        
        c_id   = str(r.get('id_interno', '')).strip()
        c_po   = str(r.get('po', '')).strip()
        c_proy = str(r.get('proyecto', '')).strip()
        c_fec  = str(r.get('fecha_llegada', '')).strip()
        c_arts = int(r.get('articulos_count', 0) or 0)
        c_req  = float(r.get('piezas_requeridas', 0) or 0)
        c_fab  = float(r.get('piezas_fabricadas', 0) or 0)
        c_ent  = float(r.get('piezas_entarimadas', 0) or 0)
        c_rem  = float(r.get('piezas_remisionadas', 0) or 0)
        c_pen  = float(r.get('piezas_pendientes', max(0.0, c_req - c_rem)) or 0)
        pct_c  = (c_rem / c_req) if c_req > 0 else 0.0
        st_txt = str(r.get('estatus_remision', 'Registrada')).strip()
        tot_val = float(r.get('total', 0) or 0)
        c_comp = str(r.get('comprador', '')).strip()
        c_sol  = str(r.get('solicitante', '')).strip()
        
        row_vals = [
            (c_id,   "center", "@", Font(name="Calibri", size=9.5, bold=True, color="0F172A"), fill_zebra),
            (c_po,   "center", "@", Font(name="Calibri", size=9.5, bold=True, color="EC2024"), fill_zebra),
            (c_proy, "left",   "@", Font(name="Calibri", size=9.5, bold=True, color="334155"), fill_zebra),
            (c_fec,  "center", "yyyy-mm-dd", Font(name="Calibri", size=9, color="64748B"), fill_zebra),
            (c_arts, "center", "#,##0", Font(name="Calibri", size=9.5, color="475569"), fill_zebra),
            (c_req,  "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="0F172A"), PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")),
            (c_fab,  "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="1D4ED8"), fill_zebra),
            (c_ent,  "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="B45309"), fill_zebra),
            (c_rem,  "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="15803D"), fill_zebra),
            (c_pen,  "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="B91C1C"), fill_zebra),
            (pct_c,  "right",  "0.0%", Font(name="Calibri", size=9.5, bold=True, color="0F172A"), fill_zebra),
            (st_txt, "center", "@", None, None), # se colorea abajo
            (tot_val,"right",  '"$"#,##0.00', Font(name="Calibri", size=9.5, bold=True, color="0F172A"), fill_zebra),
            (c_comp, "left",   "@", Font(name="Calibri", size=9, color="475569"), fill_zebra),
            (c_sol,  "left",   "@", Font(name="Calibri", size=9, color="475569"), fill_zebra),
        ]
        
        for col_i, (val, al, nf, fnt, fll) in enumerate(row_vals, start=1):
            c_cell = ws.cell(row=curr_row, column=col_i)
            c_cell.value = val
            c_cell.alignment = Alignment(horizontal=al, vertical="center")
            c_cell.number_format = nf
            c_cell.border = border_data
            if fnt: c_cell.font = fnt
            if fll: c_cell.fill = fll
            
        # Coloreo especial de badge para Estatus Entrega (Col 12)
        c_st = ws.cell(row=curr_row, column=12)
        if "Total" in st_txt or "100%" in st_txt:
            c_st.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            c_st.font = Font(name="Calibri", size=9, bold=True, color="15803D")
        elif "Parcial" in st_txt:
            c_st.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            c_st.font = Font(name="Calibri", size=9, bold=True, color="1D4ED8")
        elif "Lista" in st_txt:
            c_st.fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
            c_st.font = Font(name="Calibri", size=9, bold=True, color="6B21A8")
        elif "Fabricación" in st_txt:
            c_st.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            c_st.font = Font(name="Calibri", size=9, bold=True, color="B45309")
        else:
            c_st.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            c_st.font = Font(name="Calibri", size=9, bold=True, color="64748B")

    end_data_row = start_row + len(df_data) - 1
    
    # ── 5. Barras de Datos Nativas de Excel (Data Bars) ───────────────────────
    if end_data_row >= start_row:
        rule_fab = DataBarRule(start_type="num", start_value=0, end_type="max", color="5B9BD5", showValue=None)
        ws.conditional_formatting.add(f"G{start_row}:G{end_data_row}", rule_fab)
        
        rule_ent = DataBarRule(start_type="num", start_value=0, end_type="max", color="F59E0B", showValue=None)
        ws.conditional_formatting.add(f"H{start_row}:H{end_data_row}", rule_ent)
        
        rule_rem = DataBarRule(start_type="num", start_value=0, end_type="max", color="70AD47", showValue=None)
        ws.conditional_formatting.add(f"I{start_row}:I{end_data_row}", rule_rem)
        
        rule_pen = DataBarRule(start_type="num", start_value=0, end_type="max", color="FFC000", showValue=None)
        ws.conditional_formatting.add(f"J{start_row}:J{end_data_row}", rule_pen)
        
        rule_pct = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.0, color="6366F1", showValue=None)
        ws.conditional_formatting.add(f"K{start_row}:K{end_data_row}", rule_pct)

    # ── 6. Fila de Totales Generales ──────────────────────────────────────────
    tot_row = end_data_row + 1
    ws.row_dimensions[tot_row].height = 24
    
    ws.merge_cells(f"A{tot_row}:E{tot_row}")
    c_tot_lbl = ws[f"A{tot_row}"]
    c_tot_lbl.value = "TOTALES GENERALES CONSOLIDADOS"
    c_tot_lbl.font = Font(name="Calibri", size=10, bold=True, color=C_SLATE_DARK)
    c_tot_lbl.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fórmulas de suma nativas de Excel
    tot_cols = [
        (6,  f"=SUM(F{start_row}:F{end_data_row})", '#,##0 "pzas"', "0F172A", "F1F5F9"),
        (7,  f"=SUM(G{start_row}:G{end_data_row})", '#,##0 "pzas"', "1D4ED8", "EFF6FF"),
        (8,  f"=SUM(H{start_row}:H{end_data_row})", '#,##0 "pzas"', "B45309", "FEF3C7"),
        (9,  f"=SUM(I{start_row}:I{end_data_row})", '#,##0 "pzas"', "15803D", "DCFCE7"),
        (10, f"=SUM(J{start_row}:J{end_data_row})", '#,##0 "pzas"', "B91C1C", "FEE2E2"),
        (11, f"=I{tot_row}/F{tot_row}",            "0.0%",          "0F172A", "F1F5F9"),
        (12, "",                                    "@",             "0F172A", "F1F5F9"),
        (13, f"=SUM(M{start_row}:M{end_data_row})", '"$"#,##0.00',  "0F172A", "F1F5F9"),
        (14, "",                                    "@",             "0F172A", "F1F5F9"),
        (15, "",                                    "@",             "0F172A", "F1F5F9"),
    ]
    
    border_total = Border(
        top=Side(border_style="thin", color="0F172A"),
        bottom=Side(border_style="double", color="0F172A"), # Doble borde contable
        left=Side(border_style="thin", color="CBD5E1"),
        right=Side(border_style="thin", color="CBD5E1")
    )
    
    # Aplicar borde a A..E
    for col_i in range(1, 6):
        c = ws.cell(row=tot_row, column=col_i)
        c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c.border = border_total

    for col_i, form, nf, fg, bg in tot_cols:
        c = ws.cell(row=tot_row, column=col_i)
        if form: c.value = form
        c.number_format = nf
        c.font = Font(name="Calibri", size=10, bold=True, color=fg)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal="right" if nf != "@" else "center", vertical="center")
        c.border = border_total

    # ── 8. HOJA 2: DETALLE DE PARTIDAS (SKUs) ─────────────────────────────────
    if df_partidas is not None and not df_partidas.empty:
        ws2 = wb.create_sheet(title="Detalle_Partidas_SKU")
        ws2.views.sheetView[0].showGridLines = True
        
        # Banner hoja 2
        ws2.merge_cells("A1:K1")
        c2_t = ws2["A1"]
        c2_t.value = "INDUSTRIA SIGRAMA S.A. DE C.V.  —  DESGLOSE DE PARTIDAS (SKUs Y PRECIOS)"
        c2_t.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c2_t.fill = PatternFill(start_color=C_SLATE_DARK, end_color=C_SLATE_DARK, fill_type="solid")
        c2_t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24
        
        headers_p = [
            ("PO / Folio", 14, "center"),
            ("Item #", 8, "center"),
            ("SKU Cliente", 18, "left"),
            ("SKU Planta", 18, "left"),
            ("Descripción del Producto", 38, "left"),
            ("Cantidad Requerida", 18, "right"),
            ("Unidad", 10, "center"),
            ("Precio Unitario ($)", 18, "right"),
            ("Precio Total ($)", 18, "right"),
            ("Fecha Entrega", 14, "center"),
            ("Parcialidad / Notas", 24, "left"),
        ]
        
        ws2.row_dimensions[2].height = 24
        for col_idx, (h_name, h_w, al_h) in enumerate(headers_p, start=1):
            cell = ws2.cell(row=2, column=col_idx)
            cell.value = h_name
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=C_SLATE_MID, end_color=C_SLATE_MID, fill_type="solid")
            cell.alignment = Alignment(horizontal=al_h, vertical="center")
            cell.border = border_red_bottom
            ws2.column_dimensions[get_column_letter(col_idx)].width = h_w
            
        p_start = 3
        # Filtrar partidas solo para las POs mostradas en df_data si aplica
        valid_pos = set(df_data['po'].astype(str).str.strip().unique()) if not df_data.empty else set()
        df_part_f = df_partidas[df_partidas['po'].astype(str).str.strip().isin(valid_pos)].copy() if valid_pos else df_partidas.copy()
        
        for p_idx, (_, pr) in enumerate(df_part_f.iterrows()):
            p_row = p_start + p_idx
            ws2.row_dimensions[p_row].height = 18
            z_bg = "F8FAFC" if (p_idx % 2 == 1) else "FFFFFF"
            z_fill = PatternFill(start_color=z_bg, end_color=z_bg, fill_type="solid")
            
            p_po = str(pr.get('po', '')).strip()
            p_it = int(pr.get('item_no', 0) or 0)
            p_sk_cli = str(pr.get('sku_cliente', '')).strip()
            p_sk_pla = str(pr.get('clave_sku', '')).strip()
            p_desc = str(pr.get('descripcion_producto', '')).strip()
            p_cant = float(pr.get('cantidad_requerida', 0) or 0)
            p_unid = str(pr.get('unidad', 'PZA')).strip()
            p_pu   = float(pr.get('precio_unitario', 0) or 0)
            p_ptot = float(pr.get('precio_total', 0) or 0)
            p_fent = str(pr.get('fecha_entrega', '')).strip()
            p_parc = str(pr.get('parcialidad', pr.get('observaciones_partida', ''))).strip()
            
            row_p_vals = [
                (p_po,     "center", "@", Font(name="Calibri", size=9.5, bold=True, color="EC2024")),
                (p_it,     "center", "#,##0", None),
                (p_sk_cli, "left",   "@", Font(name="Calibri", size=9, bold=True, color="0F172A")),
                (p_sk_pla, "left",   "@", Font(name="Calibri", size=9, bold=True, color="1D4ED8")),
                (p_desc,   "left",   "@", Font(name="Calibri", size=9, color="334155")),
                (p_cant,   "right",  '#,##0 "pzas"', Font(name="Calibri", size=9.5, bold=True, color="0F172A")),
                (p_unid,   "center", "@", None),
                (p_pu,     "right",  '"$"#,##0.00', None),
                (p_ptot,   "right",  '"$"#,##0.00', Font(name="Calibri", size=9.5, bold=True, color="0F172A")),
                (p_fent,   "center", "yyyy-mm-dd", None),
                (p_parc,   "left",   "@", Font(name="Calibri", size=8.5, color="64748B")),
            ]
            
            for c_i, (p_v, p_al, p_nf, p_fnt) in enumerate(row_p_vals, start=1):
                p_c = ws2.cell(row=p_row, column=c_i)
                p_c.value = p_v
                p_c.alignment = Alignment(horizontal=p_al, vertical="center")
                p_c.number_format = p_nf
                p_c.fill = z_fill
                p_c.border = border_data
                if p_fnt: p_c.font = p_fnt
                
        p_end = p_start + len(df_part_f) - 1
        if p_end >= p_start:
            p_tot_r = p_end + 1
            ws2.row_dimensions[p_tot_r].height = 22
            ws2.merge_cells(f"A{p_tot_r}:E{p_tot_r}")
            ws2[f"A{p_tot_r}"].value = "TOTALES DE PARTIDAS"
            ws2[f"A{p_tot_r}"].font = Font(name="Calibri", size=9.5, bold=True, color=C_SLATE_DARK)
            ws2[f"A{p_tot_r}"].alignment = Alignment(horizontal="center", vertical="center")
            
            for c_i in range(1, 6):
                ws2.cell(row=p_tot_r, column=c_i).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                ws2.cell(row=p_tot_r, column=c_i).border = border_total
                
            c_cant_tot = ws2.cell(row=p_tot_r, column=6)
            c_cant_tot.value = f"=SUM(F{p_start}:F{p_end})"
            c_cant_tot.number_format = '#,##0 "pzas"'
            c_cant_tot.font = Font(name="Calibri", size=9.5, bold=True, color="0F172A")
            c_cant_tot.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            c_cant_tot.alignment = Alignment(horizontal="right", vertical="center")
            c_cant_tot.border = border_total
            
            for c_i in range(7, 9):
                c_bl = ws2.cell(row=p_tot_r, column=c_i)
                c_bl.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                c_bl.border = border_total
                
            c_monto_tot = ws2.cell(row=p_tot_r, column=9)
            c_monto_tot.value = f"=SUM(I{p_start}:I{p_end})"
            c_monto_tot.number_format = '"$"#,##0.00'
            c_monto_tot.font = Font(name="Calibri", size=9.5, bold=True, color="0F172A")
            c_monto_tot.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            c_monto_tot.alignment = Alignment(horizontal="right", vertical="center")
            c_monto_tot.border = border_total
            
            for c_i in range(10, 12):
                c_bl = ws2.cell(row=p_tot_r, column=c_i)
                c_bl.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                c_bl.border = border_total
                
            ws2.auto_filter.ref = f"A2:K{p_end}"
            
        ws2.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
