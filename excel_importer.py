import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

def generate_po_excel_template():
    """Genera la plantilla Excel oficial para la carga de Órdenes de Compra."""
    df_gen = pd.DataFrame([{
        "PO": "26083186",
        "Fecha_Pedido": "2026-08-17",
        "Proyecto": "CLOUD",
        "Solicitante": "ESEFANIA IBARRA",
        "Requisicion": "22326",
        "Destino": "ALMACEN SIGRAMA",
        "Comprador": "Josue Mesta",
        "Proveedor": "SIGRAMA PLANTA METALES",
        "Facturar_A": "INDUSTRIA SIGRAMA S.A. DE C.V.",
        "RFC": "ISI-870204-K4A",
        "Observaciones": "TAB-RQXP-MTMC ACERO PINTADO CUENTA 1881-TAB CC 736"
    }])
    
    df_part = pd.DataFrame([{
        "Item_No": 1,
        "Clave_SKU": "SWB01431",
        "Descripcion_Producto": "PP19380-03 382 X 10H BLANK DOOR",
        "Cantidad": 32.0,
        "Unidad": "PIEZA",
        "Precio_Unitario": 385.55,
        "Fecha_Entrega": "2026-08-18",
        "Parcialidad": "P1",
        "Observaciones": ""
    }])
    
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_gen.to_excel(writer, sheet_name="Datos_Generales", index=False)
        df_part.to_excel(writer, sheet_name="Partidas_PO", index=False)
        
        wb = writer.book
        fill_header = PatternFill(start_color="EC2024", end_color="EC2024", fill_type="solid")
        font_header = Font(name="Arial", color="FFFFFF", bold=True, size=11)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for sheet_name in ["Datos_Generales", "Partidas_PO"]:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align_center
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
                
    return buf.getvalue()

def parse_uploaded_excel(uploaded_file):
    """Lee un archivo Excel cargado y extrae cabecera y partidas."""
    xl = pd.ExcelFile(uploaded_file)
    
    if "Datos_Generales" not in xl.sheet_names:
        return False, "El archivo debe contener una pestaña llamada 'Datos_Generales'.", None, None
        
    df_gen = xl.parse("Datos_Generales")
    if df_gen.empty:
        return False, "La pestaña 'Datos_Generales' está vacía.", None, None
        
    row_gen = df_gen.iloc[0].to_dict()
    po_num = str(row_gen.get("PO", "")).strip()
    if not po_num or po_num.lower() == 'nan':
        return False, "No se especificó un número de 'PO' válido en 'Datos_Generales'.", None, None
        
    partidas = []
    
    # Caso 1: Tiene pestaña Partidas_PO
    if "Partidas_PO" in xl.sheet_names:
        df_part = xl.parse("Partidas_PO")
        for idx, r in df_part.iterrows():
            sku = str(r.get("Clave_SKU", r.get("SKU", ""))).strip().upper()
            if not sku or sku.lower() == 'nan':
                continue
            cant = float(r.get("Cantidad", r.get("Cantidad_Requerida", 0)) or 0)
            pu = float(r.get("Precio_Unitario", 0) or 0)
            pt = cant * pu
            f_ent = str(r.get("Fecha_Entrega", "")).strip()
            
            partidas.append({
                "item_no": int(r.get("Item_No", idx + 1)),
                "clave_sku": sku,
                "descripcion_producto": str(r.get("Descripcion_Producto", r.get("Descripcion", f"Pieza {sku}"))).strip(),
                "cantidad_requerida": cant,
                "unidad": str(r.get("Unidad", "PIEZA")).strip().upper(),
                "precio_unitario": pu,
                "precio_total": pt,
                "fecha_entrega": f_ent,
                "parcialidad": str(r.get("Parcialidad", "P1")).strip(),
                "observaciones_partida": str(r.get("Observaciones", "")).strip()
            })
            
    # Caso 2: Formato matriz tipo Detalle_Entregas
    elif "Detalle_Entregas" in xl.sheet_names:
        df_det_matrix = xl.parse("Detalle_Entregas")
        col_sku_name = df_det_matrix.columns[0]
        df_det_matrix = df_det_matrix.rename(columns={col_sku_name: 'SKU'})
        df_det_matrix['SKU'] = df_det_matrix['SKU'].astype(str).str.strip().str.upper()
        
        date_cols = df_det_matrix.columns[1:].tolist()
        df_flat = df_det_matrix.melt(id_vars=['SKU'], value_vars=date_cols, var_name='Fecha_Entrega', value_name='Cantidad_Requerida')
        df_flat['Cantidad_Requerida'] = pd.to_numeric(df_flat['Cantidad_Requerida'], errors='coerce').fillna(0)
        df_flat = df_flat[df_flat['Cantidad_Requerida'] > 0]
        
        unique_dates = sorted(df_flat['Fecha_Entrega'].unique())
        date_to_parcialidad = {date: f"P{idx+1}" for idx, date in enumerate(unique_dates)}
        
        for idx, (_, r) in enumerate(df_flat.iterrows(), start=1):
            sku = r['SKU']
            f_ent = str(r['Fecha_Entrega']).strip()
            cant = float(r['Cantidad_Requerida'])
            parc = date_to_parcialidad.get(f_ent, 'P1')
            
            partidas.append({
                "item_no": idx,
                "clave_sku": sku,
                "descripcion_producto": f"Pieza {sku}",
                "cantidad_requerida": cant,
                "unidad": "PIEZA",
                "precio_unitario": 0.0,
                "precio_total": 0.0,
                "fecha_entrega": f_ent,
                "parcialidad": parc,
                "observaciones_partida": ""
            })
            
    if not partidas:
        return False, "No se encontraron partidas válidas con cantidades mayores a 0.", None, None
        
    subtotal = sum(p["precio_total"] for p in partidas)
    iva = subtotal * 0.16
    total = subtotal + iva
    
    cabecera = {
        "po": po_num,
        "fecha_pedido": str(row_gen.get("Fecha_Pedido", "")).strip(),
        "proyecto": str(row_gen.get("Proyecto", "")).strip(),
        "solicitante": str(row_gen.get("Solicitante", "")).strip(),
        "requisicion": str(row_gen.get("Requisicion", "")).strip(),
        "destino": str(row_gen.get("Destino", "ALMACEN SIGRAMA")).strip(),
        "proveedor": str(row_gen.get("Proveedor", "SIGRAMA PLANTA METALES")).strip(),
        "proveedor_atencion": str(row_gen.get("Proveedor_Atencion", "JESUS MORALES")).strip(),
        "cliente_facturar_a": str(row_gen.get("Facturar_A", "INDUSTRIA SIGRAMA S.A. DE C.V.")).strip(),
        "cliente_rfc": str(row_gen.get("RFC", "ISI-870204-K4A")).strip(),
        "cliente_direccion": str(row_gen.get("Direccion", "C. JUAN ESCUTIA #50 COL. ABASTOS C.P. 27020 TORREON, COAH.")).strip(),
        "forma_pago": str(row_gen.get("Forma_Pago", "CONTADO / CRÉDITO")).strip(),
        "lab": str(row_gen.get("Destino", "ALMACEN SIGRAMA")).strip(),
        "tiempo_entrega": str(row_gen.get("Tiempo_Entrega", "")).strip(),
        "comprador": str(row_gen.get("Comprador", "")).strip(),
        "subtotal": subtotal,
        "descuento": 0.0,
        "iva": iva,
        "ret_iva": 0.0,
        "ret_isr": 0.0,
        "total": total,
        "moneda": "MXN",
        "observaciones": str(row_gen.get("Observaciones", "")).strip(),
        "texto_etiqueta": str(row_gen.get("Proyecto", "")).strip(),
        "color_fondo": "#EC2024",
        "color_texto": "#FFFFFF"
    }
    
    return True, f"PO {po_num} leída correctamente con {len(partidas)} partidas.", cabecera, partidas
